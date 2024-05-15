import pandas as pd
from openpyxl import Workbook, load_workbook

from connector import Connector


def main():
    print("Insert or select cashflow data.")
    print("Choose one.")
    print("1. Insert cashflow data(INSERT_FCHotel_cashflow.xlsx)")
    print("2. Select data")
    case_no = input(">>> ")

    if case_no == '1':
        insert_cashflow_data()
    elif case_no == '2':
        query = input("query > ")
        xlfilename = input("xlfilename > ")
        select_cashflow_data(query, xlfilename)
    else:
        print("Wrong Choice")


def insert_cashflow_data():
    xlfilename = 'INSERT_FCHotel_cashflow.xlsx'

    # 엑셀 파일 로드
    workbook = load_workbook(xlfilename)

    # Insert Data가 있는 Sheet 설정
    sheet = workbook['InsertSheet']

    # 첫번째 행(컬럼 이름) 가져오기 
    columns = [
        cell.value for cell 
        in next(sheet.iter_rows(min_row=1, max_row=1))
    ]

    # 데이터를 딕셔너리 리스트로 변환 
    data = [] 
    for row in sheet.iter_rows(min_row=2):
        record = {
            columns[i]: cell.value 
            for i, cell in enumerate(row)
        } 
        data.append(record)

    # 딕셔너리 리스트를 DataFrame으로 변환 
    df = pd.DataFrame(data)

    with Connector() as conn:
        # INSERT 쿼리문 작성
        query = """
            INSERT INTO cashflow_tbl (
                account_id, dw_date, category, client, 
                deposit, withdrawal, description
            )						
            VALUES (%s, %s, %s, %s, %s, %s, %s);				
        """

        # DataFrame으로부터 한줄씩 데이터를 읽어서 INSERT 쿼리문 실행
        for index, row in df.iterrows():
            val = (
                row['account_id'], row['dw_date'], row['category'], 
                row['client'], row['deposit'], row['withdrawal'], 
                row['description']
            )
            conn.cursor.execute(query, val)

        # 변경사항 커밋
        conn.connection.commit()


def select_cashflow_data(query, xlfilename):
    with Connector() as conn:
        conn.cursor.execute(query)
        result = conn.cursor.fetchall()

    # 새로운 워크북(엑셀 파일) 생성
    workbook = Workbook()

    # 활성 시트(첫번째 시트) 선택
    sheet = workbook.active

    # 데이터의 키를 엑셀 파일의 첫번째 행에 헤더로 삽입
    headers = result[0].keys()
    sheet.append(list(headers))

    # 각 딕셔너리의 값을 차례로 행에 삽입
    for item in result:
        sheet.append(list(item.values()))

    # 엑셀 파일 저장
    workbook.save(xlfilename)


if __name__ == '__main__':
    main()